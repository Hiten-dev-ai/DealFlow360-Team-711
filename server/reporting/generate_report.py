#!/usr/bin/env python3
"""Generate DealFlow360 PDF and XLSX reports from a JSON payload on stdin."""

from __future__ import annotations

import io
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


NAVY = "0B1B33"
BLUE = "1D4ED8"
PALE_BLUE = "EAF1FF"
SLATE = "516078"
LIGHT = "E5EAF2"
GREEN = "DCFCE7"
AMBER = "FEF3C7"
RED = "FEE2E2"


def safe_text(value: Any) -> str:
    text = str(value or "")
    return f"'{text}" if text[:1] in {"=", "+", "-", "@"} else text


def title_case(value: Any) -> str:
    return str(value or "").replace("_", " ").strip().title()


def indian_number(value: float) -> str:
    sign = "-" if value < 0 else ""
    whole, fraction = f"{abs(value):.2f}".split(".")
    if len(whole) > 3:
        tail = whole[-3:]
        head = whole[:-3]
        pairs = []
        while head:
            pairs.insert(0, head[-2:])
            head = head[:-2]
        whole = ",".join(pairs + [tail])
    return f"{sign}{whole}.{fraction}"


def money(minor: Any) -> str:
    return f"INR {indian_number(float(minor or 0) / 100)}"


def iso_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def status_fill(status: str) -> str:
    normalized = status.lower()
    if normalized in {"accepted", "approved", "paid", "completed"}:
        return GREEN
    if normalized in {"rejected", "expired", "overdue"}:
        return RED
    if normalized.startswith("pending") or normalized in {"negotiation", "draft"}:
        return AMBER
    return PALE_BLUE


def report_context(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], str, str, str]:
    quotes = payload.get("quotes") or []
    generated_at = iso_datetime(payload.get("generatedAt")) or datetime.now(timezone.utc).replace(tzinfo=None)
    user = payload.get("scope", {}).get("user") or "Workspace user"
    role = title_case(payload.get("scope", {}).get("role"))
    filters = payload.get("filters") or {}
    filter_parts = []
    if filters.get("status"):
        filter_parts.append(f"Status: {title_case(filters['status'])}")
    if filters.get("owner"):
        filter_parts.append(f"Owner: {filters['owner']}")
    filter_label = " | ".join(filter_parts) if filter_parts else "All visible deals"
    generated_label = generated_at.strftime("%d %b %Y, %H:%M UTC")
    return quotes, f"{safe_text(user)} - {role}", filter_label, generated_label


def generate_pdf(payload: dict[str, Any]) -> bytes:
    quotes, scope, filter_label, generated_label = report_context(payload)
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=17 * mm,
        bottomMargin=15 * mm,
        title="DealFlow360 Sales Report",
        author="DealFlow360",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=21, leading=25, textColor=colors.HexColor(f"#{NAVY}"), alignment=TA_LEFT, spaceAfter=4)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5, leading=12, textColor=colors.HexColor(f"#{SLATE}"))
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontName="Helvetica", fontSize=7.2, leading=9, textColor=colors.HexColor(f"#{NAVY}"))
    cell_right = ParagraphStyle("CellRight", parent=cell_style, alignment=TA_RIGHT)
    metric_label = ParagraphStyle("MetricLabel", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7.2, leading=9, textColor=colors.HexColor(f"#{SLATE}"))
    metric_value = ParagraphStyle("MetricValue", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=12.5, leading=15, textColor=colors.HexColor(f"#{NAVY}"))
    header_style = ParagraphStyle("Header", parent=metric_label, textColor=colors.white)

    total_minor = sum(int(quote.get("totalMinor") or 0) for quote in quotes)
    accepted = sum(1 for quote in quotes if quote.get("status") == "accepted")
    conversion = accepted / len(quotes) if quotes else 0
    average_margin = sum(float(quote.get("marginBps") or 0) for quote in quotes) / len(quotes) / 100 if quotes else 0
    story = [
        Paragraph("DealFlow360 Sales Report", title_style),
        Paragraph(f"{scope} | {filter_label} | Generated {generated_label}", subtitle_style),
        Spacer(1, 6 * mm),
    ]
    metrics = [
        [Paragraph("VISIBLE PIPELINE", metric_label), Paragraph("DEALS", metric_label), Paragraph("ACCEPTED", metric_label), Paragraph("CONVERSION", metric_label), Paragraph("AVG MARGIN", metric_label)],
        [Paragraph(money(total_minor), metric_value), Paragraph(str(len(quotes)), metric_value), Paragraph(str(accepted), metric_value), Paragraph(f"{conversion:.1%}", metric_value), Paragraph(f"{average_margin:.1f}%", metric_value)],
    ]
    metric_table = Table(metrics, colWidths=[47 * mm, 34 * mm, 34 * mm, 34 * mm, 34 * mm], rowHeights=[7 * mm, 11 * mm])
    metric_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{PALE_BLUE}")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor(f"#{LIGHT}")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{LIGHT}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.extend([metric_table, Spacer(1, 6 * mm)])

    status_counts = Counter(title_case(quote.get("status")) for quote in quotes)
    if status_counts:
        status_rows = [[Paragraph("STATUS", header_style)] + [Paragraph(status, header_style) for status in status_counts]]
        status_rows.append([Paragraph("DEALS", metric_label)] + [Paragraph(str(count), cell_style) for count in status_counts.values()])
        status_table = Table(status_rows, colWidths=[28 * mm] + [min(34 * mm, 154 * mm / len(status_counts))] * len(status_counts))
        status_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{LIGHT}")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{LIGHT}")),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([KeepTogether(status_table), Spacer(1, 6 * mm)])

    headers = ["Quotation", "Customer", "Owner", "Team", "Status", "Value", "Margin", "Risk"]
    rows = [[Paragraph(header.upper(), header_style) for header in headers]]
    for quote in quotes:
        rows.append([
            Paragraph(safe_text(quote.get("quoteNumber")), cell_style),
            Paragraph(safe_text(quote.get("customer")), cell_style),
            Paragraph(safe_text(quote.get("owner")), cell_style),
            Paragraph(safe_text(quote.get("team")), cell_style),
            Paragraph(title_case(quote.get("status")), cell_style),
            Paragraph(money(quote.get("totalMinor")), cell_right),
            Paragraph(f"{float(quote.get('marginBps') or 0) / 100:.1f}%", cell_right),
            Paragraph(str(int(quote.get("riskScore") or 0)), cell_right),
        ])
    if len(rows) == 1:
        rows.append([Paragraph("No deals match the selected filters.", cell_style)] + [""] * 7)
    deal_table = Table(rows, repeatRows=1, colWidths=[25 * mm, 43 * mm, 34 * mm, 34 * mm, 27 * mm, 34 * mm, 20 * mm, 16 * mm])
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor(f"#{LIGHT}")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor(f"#{LIGHT}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    for index in range(1, len(rows)):
        if index % 2 == 0:
            style_commands.append(("BACKGROUND", (0, index), (-1, index), colors.HexColor("#F7F9FC")))
    deal_table.setStyle(TableStyle(style_commands))
    story.append(deal_table)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor(f"#{LIGHT}"))
        canvas.line(16 * mm, 11 * mm, landscape(A4)[0] - 16 * mm, 11 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor(f"#{SLATE}"))
        canvas.drawString(16 * mm, 7 * mm, "DealFlow360 - role-scoped sales operations")
        canvas.drawRightString(landscape(A4)[0] - 16 * mm, 7 * mm, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return output.getvalue()


def generate_xlsx(payload: dict[str, Any]) -> bytes:
    quotes, scope, filter_label, generated_label = report_context(payload)
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.title = "DealFlow360 Sales Report"
    workbook.properties.subject = "Role-scoped deal pipeline"
    workbook.properties.creator = "DealFlow360"

    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.freeze_panes = "A6"
    summary["A1"] = "DealFlow360 Sales Report"
    summary["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=NAVY)
    summary["A2"] = scope
    summary["A3"] = filter_label
    summary["A4"] = f"Generated {generated_label}"
    for cell in summary["A2:A4"]:
        cell[0].font = Font(name="Aptos", size=10, color=SLATE, italic=cell[0].row == 4)

    metric_headers = ["Visible pipeline", "Deals", "Accepted", "Conversion", "Average margin"]
    for column, header in enumerate(metric_headers, 1):
        cell = summary.cell(6, column, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    data_end = max(6, 5 + len(quotes))
    if quotes:
        summary["A7"] = f"=SUM(Deals!F6:F{data_end})"
        summary["B7"] = f"=COUNTA(Deals!A6:A{data_end})"
        summary["C7"] = f'=COUNTIF(Deals!E6:E{data_end},"Accepted")'
        summary["D7"] = "=IF(B7=0,0,C7/B7)"
        summary["E7"] = f"=IF(B7=0,0,AVERAGE(Deals!G6:G{data_end}))"
    else:
        for cell in summary["A7:E7"][0]:
            cell.value = 0
    summary["A7"].number_format = '[$INR] #,##0.00'
    summary["D7"].number_format = "0.0%"
    summary["E7"].number_format = "0.0%"
    for cell in summary[7]:
        if cell.column <= 5:
            cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
            cell.font = Font(name="Aptos Display", size=15, bold=True, color=NAVY)
            cell.alignment = Alignment(horizontal="center")

    status_counts = Counter(title_case(quote.get("status")) for quote in quotes)
    summary["A10"] = "Status"
    summary["B10"] = "Deals"
    summary["C10"] = "Share"
    for cell in summary[10][:3]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    for row_index, (status, count) in enumerate(status_counts.items(), 11):
        summary.cell(row_index, 1, status)
        summary.cell(row_index, 2, count)
        summary.cell(row_index, 3, f"=IF($B$7=0,0,B{row_index}/$B$7)")
        summary.cell(row_index, 3).number_format = "0.0%"
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=status_fill(status))
    summary.column_dimensions["A"].width = 24
    for column in "BCDE":
        summary.column_dimensions[column].width = 18
    summary.row_dimensions[1].height = 30
    summary.row_dimensions[6].height = 24
    summary.row_dimensions[7].height = 31

    deals = workbook.create_sheet("Deals")
    deals.sheet_view.showGridLines = False
    deals.freeze_panes = "A6"
    deals.auto_filter.ref = f"A5:J{data_end}"
    deals["A1"] = "Deal register"
    deals["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=NAVY)
    deals["A2"] = scope
    deals["A3"] = filter_label
    deals["A2"].font = deals["A3"].font = Font(name="Aptos", size=10, color=SLATE)
    columns = ["Quotation", "Customer", "Owner", "Team", "Status", "Value (INR)", "Margin", "Risk score", "Tier", "Updated"]
    for column_index, header in enumerate(columns, 1):
        cell = deals.cell(5, column_index, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, quote in enumerate(quotes, 6):
        values = [
            safe_text(quote.get("quoteNumber")),
            safe_text(quote.get("customer")),
            safe_text(quote.get("owner")),
            safe_text(quote.get("team")),
            title_case(quote.get("status")),
            float(quote.get("totalMinor") or 0) / 100,
            float(quote.get("marginBps") or 0) / 10000,
            int(quote.get("riskScore") or 0),
            safe_text(quote.get("tier")),
            iso_datetime(quote.get("updatedAt")),
        ]
        for column_index, value in enumerate(values, 1):
            cell = deals.cell(row_index, column_index, value)
            cell.font = Font(name="Aptos", size=10, color=NAVY)
            cell.alignment = Alignment(vertical="top", wrap_text=column_index in {2, 3, 4})
        deals.cell(row_index, 5).fill = PatternFill("solid", fgColor=status_fill(values[4]))
        deals.cell(row_index, 6).number_format = '[$INR] #,##0.00'
        deals.cell(row_index, 7).number_format = "0.0%"
        deals.cell(row_index, 8).number_format = "0"
        deals.cell(row_index, 10).number_format = "dd mmm yyyy hh:mm"
        deals.row_dimensions[row_index].height = 24
    if quotes:
        table = ExcelTable(displayName="DealRegister", ref=f"A5:J{data_end}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        deals.add_table(table)
        deals.conditional_formatting.add(f"H6:H{data_end}", CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=PatternFill("solid", fgColor=RED)))
        deals.conditional_formatting.add(f"H6:H{data_end}", CellIsRule(operator="between", formula=["40", "69"], fill=PatternFill("solid", fgColor=AMBER)))
    else:
        deals["A6"] = "No deals match the selected filters."
        deals["A6"].font = Font(name="Aptos", size=10, italic=True, color=SLATE)
    widths = [16, 28, 23, 23, 20, 18, 13, 13, 14, 21]
    for column_index, width in enumerate(widths, 1):
        deals.column_dimensions[chr(64 + column_index)].width = width
    deals.row_dimensions[1].height = 28
    deals.row_dimensions[5].height = 25
    thin = Side(style="thin", color=LIGHT)
    for sheet in (summary, deals):
        used_row = sheet.max_row
        used_column = sheet.max_column
        for row in sheet.iter_rows(min_row=1, max_row=used_row, min_col=1, max_col=used_column):
            for cell in row:
                if cell.value is not None and cell.row not in {6, 7, 10}:
                    cell.border = Border(bottom=thin)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def main() -> None:
    if len(sys.argv) != 2 or sys.argv[1] not in {"pdf", "xlsx"}:
        raise ValueError("Use pdf or xlsx.")
    payload = json.load(sys.stdin)
    data = generate_pdf(payload) if sys.argv[1] == "pdf" else generate_xlsx(payload)
    sys.stdout.buffer.write(data)


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"Report generation failed: {error}", file=sys.stderr)
        raise SystemExit(1)
